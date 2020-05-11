#!/usr/bin/python3

from openpyxl import Workbook, load_workbook
from openpyxl.utils import cell
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Protection,
                             Side)
from openpyxl.comments import Comment
from configparser import ConfigParser

#COLUMN_NUMERATE = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'R',)
#SHEET_COLUMN_WIDTHS = (10, 8, 8, 39, 17, 26, 21, 22, 22, 23, 20, 21, 19, 10, 10, 21)
HEADER_FONT = Font(name='Arial',
					size=10,
					bold=True,
					italic=False,
					vertAlign=None,
					underline='none',
					strike=False,
					color='FF000000')
HEADER_FILL = PatternFill(fgColor="CCDCEC", fill_type="solid")  ## 204 220 236 (light blue)
THIN_BORDER_STYLE = Border(top=Side(border_style="thin", color="000000"),
							left=Side(border_style="thin", color="000000"),
							right=Side(border_style="thin", color="000000"),
							bottom=Side(border_style="thin", color="000000"))
TEXT_ALIGNMENT = Alignment(horizontal='left',
						vertical='top',     # {'center', 'top', 'distributed', 'justify', 'bottom'}
						text_rotation=0,
						wrap_text=True,
						shrink_to_fit=False,
						indent=0)
TEXT_FONT = Font(name='Arial',
					size=10,
					bold=False,
					italic=False)

REPORT_AUTHOR = "Ravil"

class ReportExcel():

    _row_index = 2

    def __init__(self, filename):
        # ! Check if filename exists
        # ! Load file data
        self.filename = filename
        self.wb = Workbook()
        self.sheet = self.wb.active
        self.sheet.title = "devices"
        self.wb.save(filename)

    def read_config_from_ini(self, config_filename):
        config = ConfigParser()
        print(f"Reading configuration from file '{config_filename}'")
        try:
            config.read(config_filename)
        except Exception as e:
            print(e)

    def add_sheet_header(self, column_names, column_widths, column_comments, sheet_name='devices'):
        #print(column_names)
        sheet = self.wb[sheet_name]
        for i, column in enumerate(column_names):
            sheet[cell.get_column_letter(i+1) + '1'] = column
            sheet[cell.get_column_letter(i+1) + '1'].font = HEADER_FONT
            sheet[cell.get_column_letter(i+1) + '1'].fill = HEADER_FILL
            #sheet[cell.get_column_letter(i) + '1'].border = THIN_BORDER_STYLE
            sheet.column_dimensions[cell.get_column_letter(i+1)].width = column_widths[i]
            if column_comments[i]:
                comment = Comment(column_comments[i], REPORT_AUTHOR)
                comment.width = 500
                comment.height = 50
                sheet[cell.get_column_letter(i+1) + '1'].comment = comment


        sheet.freeze_panes = sheet['A2']
        sheet.auto_filter.ref = sheet.dimensions
        self.wb.save(self.filename)

    def add_simple_sheet_header(self, column_names, sheet_name='devices'):
        DEFAULT_WIDTH = 20
        sheet = self.wb[sheet_name]
        for i, column in enumerate(column_names):
            sheet[cell.get_column_letter(i+1) + '1'] = column
            sheet[cell.get_column_letter(i+1) + '1'].font = HEADER_FONT
            sheet[cell.get_column_letter(i+1) + '1'].fill = HEADER_FILL
            sheet.column_dimensions[cell.get_column_letter(i+1)].width = DEFAULT_WIDTH

    def write_all_rows(self, devices, sheet_name='devices'):
        sheet = self.wb[sheet_name]
        row_number = 2
        for device in devices: 
            for j, value in enumerate(device):
                sheet[cell.get_column_letter(j+1) + str(row_number)] = value 
                sheet[cell.get_column_letter(j+1) + str(row_number)].font = TEXT_FONT
                #sheet[cell.get_column_letter(j) + str(row_number)].border = THIN_BORDER_STYLE
                sheet[cell.get_column_letter(j+1) + str(row_number)].alignment = TEXT_ALIGNMENT
            sheet.row_dimensions[row_number].height = 12.75
            row_number += 1
        self.wb.save(self.filename)

    @staticmethod
    def convert_value_to_multiline(value):
        try:
            return value.replace(', ', ',\n')
        except Exception:
            return value

    def write_row(self, row, sheet_name='devices', multiline=True):
        sheet = self.wb[sheet_name]
        for j, value in enumerate(row):
            value = ReportExcel.convert_value_to_multiline(value) if multiline else value
            sheet[cell.get_column_letter(j+1) + str(self._row_index)] = value 
            sheet[cell.get_column_letter(j+1) + str(self._row_index)].font = TEXT_FONT
            #sheet[cell.get_column_letter(j) + str(row_number)].border = THIN_BORDER_STYLE
            sheet[cell.get_column_letter(j+1) + str(self._row_index)].alignment = TEXT_ALIGNMENT
        sheet.row_dimensions[self._row_index].height = 12.75
        self._row_index += 1

    def save(self):
        self.wb.save(self.filename)